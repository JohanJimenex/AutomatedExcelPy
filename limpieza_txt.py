# limpieza_txt.py

# ======== IMPORTACIÓN DE MÓDULOS ========
import tkinter as tk  # Librería estándar para crear interfaces gráficas (GUI)
from tkinter import filedialog, ttk, messagebox  # Submódulos para selección de archivos, widgets avanzados y mensajes
import pandas as pd  # Manejo de datos tabulares, similar a DataTable o arrays de objetos
import sys # Acceso a parámetros y funciones del intérprete de Python
import os # Manejo de rutas y archivos del sistema operativo
import threading  # Permite ejecutar tareas en segundo plano (evita que la GUI se congele)
from openpyxl import load_workbook  # Manejo de archivos Excel
from openpyxl.worksheet.table import Table, TableStyleInfo  # Estilo y formato para tablas en Excel

cancelado = False  # Flag global para detener el procesamiento si el usuario lo desea


# ======== FUNCIONES ========

# Devuelve la ruta absoluta de un recurso (icono, imagen) compatible con PyInstaller
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS  # Ruta temporal de PyInstaller (modo empaquetado)
    except Exception:
        base_path = os.path.abspath(".")  # Modo desarrollo
    return os.path.join(base_path, relative_path)


# Marca el proceso como cancelado y actualiza la interfaz
def cancelar_proceso():
    global cancelado
    cancelado = True
    estado_label.config(text="Cancelado.")
    barra.stop()
    barra.pack_forget()
    boton_cancelar.pack_forget()
    boton_seleccionar.pack(pady=20)


# Ejecuta el procesamiento en un hilo separado (para no bloquear la GUI)
def procesar_archivos_async():
    threading.Thread(target=procesar_archivos).start()


# Función principal para procesar los archivos .txt
def procesar_archivos():
    global cancelado
    cancelado = False

    # Abre un diálogo para seleccionar uno o varios archivos .txt
    rutas_archivos = filedialog.askopenfilenames(
        title="Selecciona uno o varios archivos .txt",
        filetypes=[("Archivos de texto", "*.txt")]
    )

    if not rutas_archivos:
        return

    # Preparar la interfaz
    boton_seleccionar.pack_forget()
    boton_cancelar.pack(pady=20)
    barra.pack(pady=10)
    barra.start()
    estado_label.config(text="Procesando...")

    # Procesar cada archivo seleccionado
    for ruta_archivo in rutas_archivos:
        if cancelado:
            break

        # Leer todas las líneas del archivo
        with open(ruta_archivo, 'r') as f:
            lineas = f.readlines()

        datos = []

        # Extraer los datos de cada línea (basado en posiciones fijas)
        for linea in lineas:
            entidad = linea[0:4]
            centralta = linea[4:8]
            cuenta = linea[8:20]
            producto = linea[20:22]
            subproducto = linea[22:26]
            pan = linea[26:45]
            fecha_alta = linea[45:59].strip()
            bloqueo = linea[59:61]
            fecha_baja = linea[61:71]
            hora = linea[71:].strip()

            datos.append([
                entidad, centralta, cuenta, producto, subproducto, pan,
                fecha_alta, bloqueo, fecha_baja, hora
            ])

        # Crear un DataFrame de Pandas con las columnas especificadas
        columnas = [
            "ENTIDAD", "CENTALTA", "CUENTA", "PRODUCTO", "SUBPRODUCTO", "PAN",
            "FECHA ALTA", "BLOQUEO", "FECHA BAJA", "HORA"
        ]
        df = pd.DataFrame(datos, columns=columnas)

        # Convertir las fechas a formato datetime
        df["FECHA ALTA"] = pd.to_datetime(df["FECHA ALTA"].str.strip(), format="%d-%m-%Y", errors="coerce")
        df["FECHA BAJA"] = pd.to_datetime(df["FECHA BAJA"].str.strip(), format="%d-%m-%Y", errors="coerce").dt.date

        # Definir nombre y ruta del nuevo archivo Excel
        nombre_archivo = os.path.splitext(os.path.basename(ruta_archivo))[0]
        ruta_salida = os.path.join(os.path.dirname(ruta_archivo), f"{nombre_archivo}_procesado.xlsx")

        # Guardar el DataFrame como Excel
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Original", index=False)

        # Cargar el archivo para aplicar formato
        wb = load_workbook(ruta_salida)
        ws = wb["Original"]

        # Aplicar formato de fecha a las columnas de Excel
        for fila in range(2, len(df) + 2):
            ws[f"G{fila}"].number_format = "DD/MM/YYYY"  # FECHA ALTA
            ws[f"I{fila}"].number_format = "DD/MM/YYYY"  # FECHA BAJA

        # Crear una tabla de Excel visualmente formateada
        rango = f"A1:J{len(df) + 1}"
        tabla = Table(displayName="TablaDatos", ref=rango)
        estilo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        tabla.tableStyleInfo = estilo
        ws.add_table(tabla)

        wb.save(ruta_salida)

    # Finaliza el proceso en la GUI
    barra.stop()
    barra.pack_forget()
    boton_cancelar.pack_forget()
    boton_seleccionar.pack(pady=20)

    if cancelado:
        estado_label.config(text="Cancelado por el usuario.")
    else:
        estado_label.config(text="¡Completado!")
        messagebox.showinfo("Proceso finalizado", "✅ Archivos procesados correctamente.")


# ===============================
# INTERFAZ GRÁFICA (GUI)
# ===============================

ventana = tk.Tk()
ventana.title("Procesar Archivos")
ventana.geometry("300x200")
ventana.resizable(False, False)
ventana.iconbitmap(resource_path("assets/icono.ico"))  # Icono de la ventana

# Centrar la ventana en pantalla
ventana.update_idletasks()
x = (ventana.winfo_screenwidth() // 2) - (300 // 2)
y = (ventana.winfo_screenheight() // 2) - (200 // 2)
ventana.geometry(f"300x200+{x}+{y}")

# Texto descriptivo
descripcion = tk.Label(
    ventana,
    text="Convierte archivos .txt en hojas de Excel con formato automáticamente.",
    wraplength=280,
    justify="center",
    fg="#888888"
)
descripcion.pack(pady=(10, 5))

# Botón para seleccionar archivos
boton_seleccionar = ttk.Button(ventana, text="Seleccionar archivos", command=procesar_archivos_async, padding=7)
boton_seleccionar.pack(pady=20)

# Barra de progreso indeterminada
barra = ttk.Progressbar(ventana, mode="indeterminate")

# Etiqueta de estado ("Procesando...", "Cancelado", etc.)
estado_label = tk.Label(ventana, text="", fg="#555555")
estado_label.pack()

# Botón para cancelar el proceso
boton_cancelar = ttk.Button(ventana, text="Cancelar", command=cancelar_proceso)

# Inicia la ventana principal
ventana.mainloop()
